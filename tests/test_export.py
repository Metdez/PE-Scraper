from __future__ import annotations

import csv as csv_mod

from openpyxl import load_workbook

from pescraper import db
from pescraper.export import export_csv, export_excel
from pescraper.models import FirmRecord, FirmStatus


def _seeded_conn(tmp_path):
    conn = db.connect(db.init_db(tmp_path / "pipeline.db"))
    db.upsert_firm(
        conn,
        FirmRecord(
            firm_name="Acme Capital",
            website="https://acme.example",
            ebitda_min_musd=5.0,
            confidence=0.8,
            needs_review=False,
            status=FirmStatus.COMPLETE,
        ),
    )
    db.upsert_firm(
        conn,
        FirmRecord(
            firm_name="Blocked Firm",
            website="https://blocked.example",
            confidence=0.05,
            needs_review=True,
            status=FirmStatus.NEEDS_REVIEW,
        ),
    )
    return conn


def test_export_csv_writes_header_and_rows(tmp_path) -> None:
    conn = _seeded_conn(tmp_path)
    path = export_csv(conn, tmp_path / "out" / "firms.csv")
    assert path.exists()
    with open(path, encoding="utf-8") as fh:
        rows = list(csv_mod.reader(fh))
    assert rows[0][0] == "Firm Name"
    assert len(rows) == 3  # header + 2 firms
    names = {r[0] for r in rows[1:]}
    assert names == {"Acme Capital", "Blocked Firm"}


def test_export_excel_writes_firms_and_summary_sheets(tmp_path) -> None:
    conn = _seeded_conn(tmp_path)
    path = export_excel(conn, tmp_path / "out" / "firms.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    assert "Firms" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    firms_ws = wb["Firms"]
    assert firms_ws.cell(row=1, column=1).value == "Firm Name"
    assert firms_ws.max_row == 3
