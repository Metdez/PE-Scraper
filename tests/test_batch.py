from __future__ import annotations

import csv

from openpyxl import load_workbook

from pescraper import db
from pescraper.models import FirmRecord


def _conn(tmp_path):
    path = tmp_path / "pipeline.db"
    db.init_db(path)
    return db.connect(path)


def test_claim_next_job_honors_priority_then_fifo(tmp_path) -> None:
    from pescraper.queue import claim_next_job, enqueue

    conn = _conn(tmp_path)
    enqueue(conn, "https://batch.example", priority=9)
    urgent = enqueue(conn, "https://urgent.example", priority=0)

    claimed = claim_next_job(conn)

    assert claimed is not None
    assert claimed.id == urgent
    assert claimed.website == "https://urgent.example"


def test_run_batch_records_failure_and_continues(tmp_path) -> None:
    from pescraper.queue import enqueue, queue_summary
    from pescraper.worker import run_batch

    conn = _conn(tmp_path)
    enqueue(conn, "https://bad.example")
    enqueue(conn, "https://good.example")

    def processor(url: str):
        if "bad" in url:
            raise RuntimeError("blocked")
        return FirmRecord(firm_name="Good Capital", website=url)

    result = run_batch(conn, processor, limit=2)

    assert result.completed == 1
    assert result.failed == 1
    assert queue_summary(conn) == {"complete": 1, "failed": 1}
    assert db.get_firm(conn, "https://good.example") is not None


def test_export_dataset_writes_csv_and_excel_summary(tmp_path) -> None:
    from pescraper.exporter import export_dataset

    conn = _conn(tmp_path)
    db.upsert_firm(
        conn,
        FirmRecord(
            firm_name="Acme Capital",
            website="https://acme.example",
            ebitda_min_musd=5,
            needs_review=True,
        ),
    )

    csv_path, xlsx_path = export_dataset(conn, tmp_path / "firms")

    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["firm_name"] == "Acme Capital"

    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["Firms", "Summary"]
    assert workbook["Summary"]["B2"].value == 1
