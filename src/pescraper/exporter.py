"""CSV and styled Excel export for the 24-column firm dataset."""

from __future__ import annotations

import csv
import sqlite3
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from pescraper.models import FIRM_COLUMNS


def export_dataset(conn: sqlite3.Connection, output: str | Path) -> tuple[Path, Path]:
    base = Path(output)
    base.parent.mkdir(parents=True, exist_ok=True)
    csv_path = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")
    rows = [dict(row) for row in conn.execute("SELECT * FROM firms ORDER BY firm_name")]

    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(FIRM_COLUMNS), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    firms = workbook.active
    firms.title = "Firms"
    firms.append(list(FIRM_COLUMNS))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in firms[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for row in rows:
        firms.append([row.get(column) for column in FIRM_COLUMNS])
        if row.get("needs_review"):
            for cell in firms[firms.max_row]:
                cell.fill = review_fill
    firms.freeze_panes = "A2"
    firms.auto_filter.ref = firms.dimensions

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Count"])
    summary.append(["Total firms", len(rows)])
    statuses = Counter(str(row.get("status")) for row in rows)
    for status, count in sorted(statuses.items()):
        summary.append([status, count])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    workbook.save(xlsx_path)
    return csv_path, xlsx_path


__all__ = ["export_dataset"]
