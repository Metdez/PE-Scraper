"""Excel/CSV export of the firms dataset. The store (SQLite) is the source of
truth; export is a read-only view over it — never the other way around.
"""

from __future__ import annotations

import csv as csv_mod
from pathlib import Path

from pescraper.db import all_firms, firm_status_summary
from pescraper.models import FIRM_COLUMNS, FirmRecord

DISPLAY_LABELS: dict[str, str] = {
    "firm_name": "Firm Name",
    "type": "Type",
    "state": "State",
    "city": "City",
    "website": "Website",
    "us_investments": "US Investments",
    "rev_min_musd": "Rev Min ($M)",
    "rev_max_musd": "Rev Max ($M)",
    "ebitda_min_musd": "EBITDA Min ($M)",
    "ebitda_max_musd": "EBITDA Max ($M)",
    "ev_min_musd": "EV Min ($M)",
    "ev_max_musd": "EV Max ($M)",
    "check_min_musd": "Check Min ($M)",
    "check_max_musd": "Check Max ($M)",
    "deal_types": "Deal Types",
    "sector_tier1": "Sector Tier 1",
    "aum_musd": "AUM ($M)",
    "activity": "Activity",
    "last_deal": "Last Deal",
    "fund_name": "Fund Name",
    "confidence": "Confidence",
    "needs_review": "Needs Review",
    "last_checked": "Last Checked",
    "status": "Status",
}

HEADER_ROW = [DISPLAY_LABELS[c] for c in FIRM_COLUMNS]


def _row_values(record: FirmRecord) -> list:
    data = record.model_dump()
    data["status"] = record.status.value
    return [data[c] for c in FIRM_COLUMNS]


def export_csv(conn, path: str | Path) -> Path:
    """Write every firm row to a CSV file. Returns the path written."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    records = all_firms(conn)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv_mod.writer(fh)
        writer.writerow(HEADER_ROW)
        for record in records:
            writer.writerow(_row_values(record))
    return path


def export_excel(conn, path: str | Path) -> Path:
    """Write a color-coded .xlsx workbook (Firms sheet + Summary sheet)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    records = all_firms(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Firms"
    ws.append(HEADER_ROW)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    confidence_col = FIRM_COLUMNS.index("confidence") + 1
    needs_review_col = FIRM_COLUMNS.index("needs_review") + 1

    for record in records:
        ws.append(_row_values(record))
        row_idx = ws.max_row
        confidence = record.confidence or 0.0
        fill = green if confidence >= 0.7 else yellow if confidence >= 0.3 else red
        ws.cell(row=row_idx, column=confidence_col).fill = fill
        if record.needs_review:
            ws.cell(row=row_idx, column=needs_review_col).fill = red

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Status", "Count"])
    for status, count in sorted(firm_status_summary(conn).items()):
        summary_ws.append([status, count])
    summary_ws.append([])
    summary_ws.append(["Total firms", len(records)])
    if records:
        avg_conf = sum(r.confidence or 0.0 for r in records) / len(records)
        summary_ws.append(["Average confidence", round(avg_conf, 3)])

    wb.save(path)
    return path


__all__ = ["DISPLAY_LABELS", "HEADER_ROW", "export_csv", "export_excel"]
